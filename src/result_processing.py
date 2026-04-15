import os
import pandas as pd
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Создаём df с результатами детекции и распознавания по каждому изображению
def create_result_df(prefix, threshold, df):
    df['text'] = df['text'].astype('str')
    df['prefix'] = str(prefix)
    df['confidence'] = df['detection_confidence'] * df['recognition_confidence']
    df['confidence'] = df['confidence'].round(3)
    df['recognition_confidence'] = df['recognition_confidence'].round(3)
    df['new_name1'] = df.apply(
        lambda row: '!_' + row['file_name'] if row['detection_confidence'] == 0 or row['text'] == 'no text'
        else '!_' + row['prefix'] + row['text'].zfill(4) + row['file_ext'] if row['confidence'] < threshold
        else row['prefix'] + row['text'].zfill(4) + row['file_ext'],
        axis=1
    )
    df['new_name1'] = df.apply(
        lambda row: row['new_name1'] if row['new_name1'].replace(row['file_ext'],'').isdigit() or row['detection_confidence'] == 0
        else row['new_name1'].zfill(5+len(row['file_ext'])),
        axis=1
    )
    df['name_iter'] = 0
    mask = (df['text'] != 'no detection') & (df['text'] != 'no text')
    df.loc[mask, 'name_iter'] = df.loc[mask].groupby('text').cumcount()
    df['new_name'] = df.apply(
        lambda row: row['new_name1'] if row['name_iter'] == 0
        else str(row['new_name1']).replace(str(row['file_ext']),'') + '_' + str(row['name_iter']) + row['file_ext'],
        axis=1
    )
    return df

# Рассчитываем новое имя файла
def rename_files_from_dataframe(folder_path, df, old_name_col='file_name', new_name_col='new_name'):
    results = {
        'success': [],
        'not_found': [],
        'errors': []
    }
    # Проверяем существование папки
    if not os.path.exists(folder_path):
        raise FileNotFoundError(f"Папка не найдена: {folder_path}")

    # Получаем список файлов в папке
    existing_files = set(os.listdir(folder_path))

    for _, row in df.iterrows():
        old_name = row[old_name_col]
        new_name = row[new_name_col]

        # Проверяем, существует ли файл
        if old_name not in existing_files:
            results['not_found'].append(old_name)
            continue

        try:
            # Формируем полные пути
            old_path = os.path.join(folder_path, old_name)
            new_path = os.path.join(folder_path, new_name)

            # Переименовываем файл
            if old_path != new_path:
                os.rename(old_path, new_path)
                results['success'].append((old_name, new_name))

        except Exception as e:
            results['errors'].append(f"{old_name} -> {new_name}: {str(e)}")

    return results


# Печатаем свод по результатам в xls
def print_to_xls(df, threshold, folder_path):

    cnt_no_detections = df.query('detection_confidence==0').shape[0]
    cnt_low_confidence = df.query('confidence < @threshold and detection_confidence!=0').shape[0]
    cnt_renamed = df.query('confidence >= @threshold').shape[0]

    with pd.ExcelWriter(folder_path + '/logs_test.xlsx', engine='openpyxl') as writer:
        df_cut = df[['file_name','detection_confidence','recognition_confidence','confidence','text', 'new_name']]
        df_cut.to_excel(writer, sheet_name='Отчёт', index=False, startrow=3)

        # Получаем лист
        worksheet = writer.sheets['Отчёт']

        # Определяем стиль для заголовка

        header_font = Font(bold=True, color='000000')  # Чёрный шрифт
        header_fill = PatternFill(
            start_color='ADD8E6',  # Светло‑синий цвет (Light Blue)
            end_color='ADD8E6',
            fill_type='solid'
        )
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Добавляем расчётные показатели в первые три строки
        worksheet['A1'] = f'Переименованных файлов: {cnt_renamed}'
        worksheet['A2'] = f'Переименованных с низкой надёжностью распознавания: {cnt_low_confidence}'
        worksheet['A3'] = f'Не переименованных: {cnt_no_detections}'

        # Стилизуем строки с показателями (жирный чёрный шрифт)
        for row in range(1, 4):
            worksheet.cell(row=row, column=1).font = Font(bold=True, color='000000')

        # Применяем стиль к заголовкам (предполагаем, что заголовки в первой строке)
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=4, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        # Устанавливаем ширину столбцов (50 символов)
        for col in range(1, worksheet.max_column + 1):
            column_letter = get_column_letter(col)
            worksheet.column_dimensions[column_letter].width = 25

# Выгружаем результаты распознавания в csv
def print_to_csv(df, folder_path):
    df_cut = df[['file_name', 'text']]
    df_cut.to_csv(folder_path + '/recognition_results.csv', encoding='utf-8', index=False)
